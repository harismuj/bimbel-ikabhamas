from flask import Blueprint, render_template, url_for, flash, redirect, request, jsonify, send_file
from flask_login import current_user, login_required, logout_user
from app.models import SocialMediaLink, User, Layanan, Alumni, ProgramFasilitas, Keunggulan, Artikel, Visitor, Profil, Jasmani
from app.forms import ProfileForm, JasmaniForm, SocMedForm
from app import db
import os
import io
from datetime import datetime
from app.libmain import calc_interpolasi, calc_run, calc_up
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

admin = Blueprint('admin', __name__)

def save_logo_file(file):
    filename = secure_filename(file.filename)
    name, ext = os.path.splitext(filename)
    current_date = datetime.now().strftime("%d-%m-%Y")
    new_filename = f"logo-{current_date}{ext}"
    file_path = os.path.join(os.getcwd(), 'app', 'static', 'img', new_filename)
    file.save(file_path)
    return new_filename

@admin.context_processor
def inject_data():
    profiles = Profil.query.all()
    articles = Artikel.query.all()
    programs = ProgramFasilitas.query.all()
    return dict(profiles=profiles, articles=articles, programs=programs)


@admin.route('/export')
def export():
    tanggal_mulai = request.args.get('tanggal_mulai')
    tanggal_akhir = request.args.get('tanggal_akhir')

    if tanggal_mulai and tanggal_akhir:
        try:
            tanggal_mulai_dt = datetime.strptime(tanggal_mulai, '%Y-%m-%d')
            tanggal_akhir_dt = datetime.strptime(tanggal_akhir, '%Y-%m-%d')
        except ValueError:
            return '''<script>
                    alert('Tanggal Belum dimasukkan!,Pilih rentang tanggal dan klik filter jika ingin dicetak');
                    window.location.href = "dashboard/jasmani";
                    </script>''', 400
        
        jasmanis = Jasmani.query.filter(Jasmani.timestamp >= tanggal_mulai_dt, Jasmani.timestamp <= tanggal_akhir_dt).order_by(Jasmani.nilai_akhir.desc()).all()
    else:
        jasmanis = Jasmani.query.order_by(Jasmani.nilai_akhir.desc()).all()

    data = []
    i=0
    for jasmani in jasmanis:
        i += 1
        data.append({
            'No.':i,
            'Nama Siswa': jasmani.nama_siswa,
            'Jenis Kelamin': jasmani.jenis_kelamin,
            'Lari': jasmani.nilai_lari,
            'Pull-Up/Chinning': jasmani.nilai_pullup,
            'Sit-Up': jasmani.nilai_situp,
            'Push-Up': jasmani.nilai_pushup,
            'Shuttle Run': jasmani.nilai_shutrun,
            'Renang': jasmani.nilai_renang,
            'Nilai Akhir': jasmani.nilai_akhir,
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()

    # Menggunakan openpyxl untuk menambahkan format
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Data Jasmani"

     # Menambahkan header tambahan
    ws.merge_cells('A1:J1')
    ws['A1'] = "Laporan Data Jasmani"
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:J2')
    if tanggal_mulai and tanggal_akhir:
        ws['A2'] = f"Periode: {tanggal_mulai} sampai {tanggal_akhir}"
    else:
        ws['A2'] = "Periode: Semua Data"
    ws['A2'].font = Font(size=13)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Menambahkan data ke worksheet mulai dari baris 5
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 5):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Membuat border
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if r_idx == 5:  # Header row
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = 15  # Set column width


    # Save the workbook to the output stream
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="LaporanJasmani_"+tanggal_mulai+".xlsx", as_attachment=True)

@admin.route("/dashboard")
@login_required
def dashboard():
    unique_ips_count = Visitor.query.distinct(Visitor.ip_address).count()
    return render_template('admin/dashboard.html', title='Dashboard', unique_ips_count=unique_ips_count)

@admin.route("/dashboard/layanan")
@login_required
def admin_layanan():
    services = Layanan.query.all()
    return render_template('admin/admin_layanan.html',title='layanan', services=services )

@admin.route("/dashboard/alumni")
@login_required
def admin_alumni():
    alumni_list = Alumni.query.all()
    return render_template('admin/admin_alumni.html',title='alumni', alumni_list=alumni_list)

@admin.route("/dashboard/program")
@login_required
def admin_program():
    programs = ProgramFasilitas.query.all()
    return render_template('admin/admin_program.html',title='Program', programs=programs)

@admin.route("/dashboard/keunggulan")
@login_required
def admin_keunggulan():
    keunggulan = Keunggulan.query.all()
    return render_template('admin/admin_keunggulan.html',title='Keunggulan', keunggulan = keunggulan)

@admin.route("/dashboard/artikel")
@login_required
def admin_artikel():
    articles = Artikel.query.all()
    return render_template('admin/admin_artikel.html',title='Artikel', articles=articles)

@admin.route("/dashboard/jasmani", methods=['GET', 'POST'])
@login_required
def admin_jasmani():
    form = JasmaniForm()

    if form.validate_on_submit():
        # Ambil data dari formulir
        nama_siswa = form.nama_siswa.data
        jenis_kelamin = form.jenis_kelamin.data
        nilai_lari = form.nilai_lari.data
        nilai_pullup = form.nilai_pullup.data
        nilai_situp = form.nilai_situp.data
        nilai_pushup = form.nilai_pushup.data
        nilai_shutrun = form.nilai_shutrun.data
        nilai_renang = form.nilai_renang.data
        if jenis_kelamin == 'Laki-laki':
            fn_lari = calc_run(nilai_lari,3500)
            fn_pullup = calc_up(nilai_pullup, 17)
            fn_situp = calc_up(nilai_situp, 42)
            fn_pushup = calc_up(nilai_pushup, 42)
            fn_shutrun = calc_interpolasi(nilai_shutrun, [(16.2, 100), (17, 90), (19, 51)])
            fn_renang = calc_interpolasi(nilai_renang,[(14, 100), (20, 91), (26, 82),(48,51)])
       
        elif jenis_kelamin == 'Perempuan':
            fn_lari = calc_run(nilai_lari,3100)
            fn_pullup = calc_up(nilai_pullup, 72)
            fn_situp = calc_up(nilai_situp, 50)
            fn_pushup = calc_up(nilai_pushup, 37)
            fn_shutrun = calc_interpolasi(nilai_shutrun, [(17.6, 100),(18.6, 90),(20, 76),(22.5, 51)])
            fn_renang = calc_interpolasi(nilai_renang,[(20,100),(30,85),(40,70),(50,55)])
        nilai_akhir = int((70/100*(fn_lari+fn_pullup+fn_pushup+fn_situp+fn_shutrun))/5)+(30/100*fn_renang)
     
        # Simpan data ke tabel Jasmani
        new_jasmani = Jasmani(
            nama_siswa=nama_siswa,
            jenis_kelamin=jenis_kelamin,
            nilai_lari=fn_lari,
            nilai_pullup=fn_pullup,
            nilai_situp=fn_situp,
            nilai_pushup=fn_pushup,
            nilai_shutrun=fn_shutrun,
            nilai_renang=fn_renang,
            nilai_akhir=nilai_akhir
        )
        db.session.add(new_jasmani)
        db.session.commit()

        # Berikan pesan flash
        flash('Data jasmani berhasil disimpan!',category="success")

        return redirect(url_for('admin.admin_jasmani'))
    tanggal_mulai = request.args.get('tanggal_mulai')
    tanggal_akhir = request.args.get('tanggal_akhir')

    if tanggal_mulai and tanggal_akhir:
        jasmanis = Jasmani.query.filter(Jasmani.timestamp >= tanggal_mulai, Jasmani.timestamp <= tanggal_akhir).order_by(Jasmani.nilai_akhir.desc()).all()
    else:
        jasmanis = Jasmani.query.order_by(Jasmani.nilai_akhir.desc()).all()
    return render_template('admin/admin_jasmani.html',title='Jasmani', form=form, jasmanis=jasmanis)

@admin.route('/dashboard/profile', methods=['GET', 'POST'])
@login_required
def admin_profil(profile_id=1):
    profile = Profil.query.get_or_404(profile_id)
    form = ProfileForm(obj=profile)

    if form.validate_on_submit():
        form.populate_obj(profile)

        if form.logo.data:
            file_path = save_logo_file(form.logo.data)
            profile.logo = file_path

        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('admin.admin_profil', profile_id=profile.id))
    
    return render_template('admin/admin_profil.html', title='Profile', form=form, profile_id=profile_id, profile=profile)

@admin.route('/dashboard/socmed',methods=['GET','POST'])
@login_required
def admin_socmed(socmed_id=1):
    socmed = SocialMediaLink.query.get_or_404(socmed_id)
    form = SocMedForm(obj=socmed)
    if form.validate_on_submit():
        form.populate_obj(socmed)
        db.session.commit()
        flash('Socmed Links updated successfully!', 'success')
        return redirect(url_for('admin.admin_socmed', socmed_id=socmed.id))
    
    return render_template('admin/socmed.html', title='Social Media Links', form=form, socmed_id=socmed_id, socmed=socmed)

@admin.route("/dashboard/jasmani2", methods=['GET', 'POST'])
@login_required
def admin_jasmani2():
    form = JasmaniForm()

    if form.validate_on_submit():
        # Ambil data dari formulir
        nama_siswa = form.nama_siswa.data
        jenis_kelamin = form.jenis_kelamin.data
        nilai_lari = form.nilai_lari.data
        nilai_pullup = form.nilai_pullup.data
        nilai_situp = form.nilai_situp.data
        nilai_pushup = form.nilai_pushup.data
        nilai_shutrun = form.nilai_shutrun.data
        nilai_renang = form.nilai_renang.data
        if jenis_kelamin == 'Laki-laki':
            fn_lari = calc_run(nilai_lari,3500)
            fn_pullup = calc_up(nilai_pullup, 17)
            fn_situp = calc_up(nilai_situp, 42)
            fn_pushup = calc_up(nilai_pushup, 42)
            fn_shutrun = calc_interpolasi(nilai_shutrun, [(16.2, 100), (17, 90), (19, 51)])
            fn_renang = calc_interpolasi(nilai_renang,[(14, 100), (20, 91), (26, 82),(48,51)])
       
        elif jenis_kelamin == 'Perempuan':
            fn_lari = calc_run(nilai_lari,3100)
            fn_pullup = calc_up(nilai_pullup, 72)
            fn_situp = calc_up(nilai_situp, 50)
            fn_pushup = calc_up(nilai_pushup, 37)
            fn_shutrun = calc_interpolasi(nilai_shutrun, [(17.6, 100),(18.6, 90),(20, 76),(22.5, 51)])
            fn_renang = calc_interpolasi(nilai_renang,[(20,100),(30,85),(40,70),(50,55)])
        nilai_akhir = int((70/100*(fn_lari+fn_pullup+fn_pushup+fn_situp+fn_shutrun))/5)+(30/100*fn_renang)
     
        # Simpan data ke tabel Jasmani
        new_jasmani = Jasmani(
            nama_siswa=nama_siswa,
            jenis_kelamin=jenis_kelamin,
            nilai_lari=fn_lari,
            nilai_pullup=fn_pullup,
            nilai_situp=fn_situp,
            nilai_pushup=fn_pushup,
            nilai_shutrun=fn_shutrun,
            nilai_renang=fn_renang,
            nilai_akhir=nilai_akhir
        )
        db.session.add(new_jasmani)
        db.session.commit()

        # Berikan pesan flash
        flash('Data jasmani berhasil disimpan!',category="success")

        return redirect(url_for('admin.admin_jasmani'))
    tanggal_mulai = request.args.get('tanggal_mulai')
    tanggal_akhir = request.args.get('tanggal_akhir')

    if tanggal_mulai and tanggal_akhir:
        jasmanis = Jasmani.query.filter(Jasmani.timestamp >= tanggal_mulai, Jasmani.timestamp <= tanggal_akhir).order_by(Jasmani.nilai_akhir.desc()).all()
    else:
        jasmanis = Jasmani.query.order_by(Jasmani.nilai_akhir.desc()).all()
    return render_template('admin/admin_jasmani_desk.html',title='Jasmani', form=form, jasmanis=jasmanis)